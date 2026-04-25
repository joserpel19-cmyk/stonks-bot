# =============================================================
#  STONKS BOT - GENERADOR DE EXCEL CON GRÁFICAS
# -------------------------------------------------------------
#  Lee pending_bets.csv, settled_bets.csv y bankroll_state.json
#  y genera STONKS_Bot_Dashboard.xlsx con 6 hojas + gráficas.
#
#  Se ejecuta automáticamente al final de 03_paper_trading.py.
#  También se puede lanzar a mano:
#       python 06_generar_excel.py
# =============================================================

import json, csv, os
import datetime as dt
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
except ImportError:
    raise SystemExit(
        "Falta openpyxl. Instálalo con:\n"
        "    pip install openpyxl"
    )

DIR = Path(__file__).parent
OUT = DIR / "STONKS_Bot_Dashboard.xlsx"
BANKROLL_INICIAL = 20.00


def leer_json(nombre, default):
    p = DIR / nombre
    if not p.exists(): return default
    try:
        raw = p.read_text(encoding="utf-8").strip()
        if not raw: return default
        return json.loads(raw)
    except Exception:
        return default


def leer_csv(nombre):
    p = DIR / nombre
    if not p.exists(): return []
    with open(p, "r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def generar():
    estado  = leer_json("bankroll_state.json", {"bankroll": BANKROLL_INICIAL, "inicio": dt.datetime.now().isoformat(), "historial": []})
    pending = leer_csv("pending_bets.csv")
    settled = leer_csv("settled_bets.csv")

    wb = Workbook()
    FONT = "Arial"
    HEADER_FILL = PatternFill("solid", start_color="1F3A5F")
    HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=16)
    TITLE_FILL = PatternFill("solid", start_color="0F1F33")
    SUB_FONT = Font(name=FONT, color="777777", size=10, italic=True)
    BODY_FONT = Font(name=FONT, size=10)
    KPI_LBL_FONT = Font(name=FONT, size=9, color="666666", bold=True)
    thin = Side(border_style="thin", color="D0D0D0")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    bankroll = float(estado.get("bankroll", BANKROLL_INICIAL))
    n_set = len(settled)
    ganadas = sum(1 for s in settled if s.get("resultado") == "GANADA")
    perdidas = n_set - ganadas
    staked = sum(float(s.get("stake_eur", 0) or 0) for s in settled)
    exposicion = sum(float(p.get("stake_eur", 0) or 0) for p in pending)

    try:
        ini = dt.datetime.fromisoformat(estado.get("inicio"))
        dias = max((dt.datetime.now() - ini).days, 1)
    except Exception:
        dias = 1

    # ------- HOJA 1: RESUMEN -------
    ws = wb.active; ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:H3")
    ws["B2"] = "STONKS BOT  ·  Dashboard"
    ws["B2"].font = TITLE_FONT; ws["B2"].fill = TITLE_FILL
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B4:H4")
    ws["B4"] = f"Actualizado: {dt.datetime.now():%d/%m/%Y %H:%M:%S}  ·  Día {dias}  ·  Sin dinero real"
    ws["B4"].font = SUB_FONT
    ws["B4"].alignment = Alignment(horizontal="center")

    kpis = [
        ("Bankroll inicial (€)", BANKROLL_INICIAL, "0.00"),
        ("Bankroll actual (€)", bankroll, "0.00"),
        ("P&L total (€)", "=C7-C6", "+0.00;-0.00;0.00"),
        ("Volumen apostado (€)", staked, "0.00"),
        ("ROI (%)", "=IF(C9>0,C8/C9,0)", "+0.00%;-0.00%;0.00%"),
        ("Apuestas resueltas", n_set, "0"),
        ("  · Ganadas", ganadas, "0"),
        ("  · Perdidas", perdidas, "0"),
        ("% Acierto", "=IF(C11>0,C12/C11,0)", "0.00%"),
        ("Apuestas abiertas", len(pending), "0"),
        ("Exposición actual (€)", exposicion, "0.00"),
        ("Días de paper trading", dias, "0"),
    ]
    for i, (lbl, val, fmt) in enumerate(kpis, start=6):
        ws.cell(row=i, column=2, value=lbl).font = KPI_LBL_FONT
        c = ws.cell(row=i, column=3, value=val)
        c.number_format = fmt
        ws[f"B{i}"].alignment = Alignment(horizontal="left", indent=1)
        ws[f"C{i}"].alignment = Alignment(horizontal="right", indent=1)
        ws[f"B{i}"].border = BORDER; ws[f"C{i}"].border = BORDER

    ws["E6"] = "¿Cómo leer esta hoja?"
    ws["E6"].font = Font(name=FONT, bold=True, size=11, color="1F3A5F")
    notas = [
        "• Bankroll: dinero virtual (empiezas con 20 €).",
        "• P&L: beneficio o pérdida acumulada desde el día 1.",
        "• ROI: rentabilidad sobre el dinero apostado.",
        "• % Acierto: de las apuestas liquidadas, cuántas has ganado.",
        "• Apuestas abiertas: ya colocadas pero esperando el resultado.",
        "• Exposición: cuánto dinero está \"en juego\" en este momento.",
        "",
        "Un ROI > 0 mantenido durante 2 semanas es la señal para",
        "pasar de paper trading a dinero real en Betfair.",
    ]
    for i, t in enumerate(notas, start=7):
        ws.cell(row=i, column=5, value=t).font = Font(name=FONT, size=10, color="444444")
        ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=8)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 2
    for col in "EFGH": ws.column_dimensions[col].width = 18

    # ------- HOJA 2: APUESTAS ABIERTAS -------
    ws2 = wb.create_sheet("Apuestas abiertas")
    ws2.sheet_view.showGridLines = False
    headers = ["Fecha","Deporte","Partido","Selección","Casa","Cuota","Prob. real","Edge %","Stake (€)","Ganancia si acierta (€)"]
    for j, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center"); c.border = BORDER

    row = 2
    for p in pending:
        try: cuota = float(p.get("cuota", 0))
        except: cuota = 0
        try: stake = float(p.get("stake_eur", 0))
        except: stake = 0
        try: prob  = float(p.get("prob_real", 0))
        except: prob = 0
        try: edge  = float(p.get("edge_pct", 0)) / 100.0
        except: edge = 0
        fecha = p.get("fecha", "")[:16].replace("T"," ")
        ws2.cell(row=row, column=1, value=fecha)
        ws2.cell(row=row, column=2, value=p.get("deporte","").replace("_"," "))
        ws2.cell(row=row, column=3, value=p.get("partido",""))
        ws2.cell(row=row, column=4, value=p.get("seleccion",""))
        ws2.cell(row=row, column=5, value=p.get("casa",""))
        ws2.cell(row=row, column=6, value=cuota).number_format = "0.00"
        ws2.cell(row=row, column=7, value=prob).number_format = "0.00%"
        ws2.cell(row=row, column=8, value=edge).number_format = "+0.00%;-0.00%;0.00%"
        ws2.cell(row=row, column=9, value=stake).number_format = "0.00"
        ws2.cell(row=row, column=10, value=f"=F{row}*I{row}-I{row}").number_format = "0.00"
        for col in range(1, 11):
            ws2.cell(row=row, column=col).font = BODY_FONT
            ws2.cell(row=row, column=col).border = BORDER
        row += 1

    ws2.freeze_panes = "A2"
    if row > 2:
        ws2.auto_filter.ref = f"A1:J{row-1}"
        ws2.conditional_formatting.add(
            f"H2:H{row-1}",
            ColorScaleRule(start_type="min", start_color="FFEB84",
                           mid_type="percentile", mid_value=50, mid_color="63BE7B",
                           end_type="max", end_color="1E8449")
        )
    widths = [18,18,38,26,16,8,12,12,12,20]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ------- HOJA 3: APUESTAS LIQUIDADAS -------
    ws3 = wb.create_sheet("Apuestas liquidadas")
    ws3.sheet_view.showGridLines = False
    heads3 = ["Fecha","Deporte","Partido","Selección","Casa","Cuota","Stake (€)","Resultado","P&L (€)"]
    for j, h in enumerate(heads3, start=1):
        c = ws3.cell(row=1, column=j, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center"); c.border = BORDER

    row = 2
    for s in settled:
        try: cuota = float(s.get("cuota", 0))
        except: cuota = 0
        try: stake = float(s.get("stake_eur", 0))
        except: stake = 0
        try: pnl_s = float(s.get("pnl", 0))
        except: pnl_s = 0
        fecha = s.get("fecha", "")[:16].replace("T"," ")
        ws3.cell(row=row, column=1, value=fecha)
        ws3.cell(row=row, column=2, value=s.get("deporte","").replace("_"," "))
        ws3.cell(row=row, column=3, value=s.get("partido",""))
        ws3.cell(row=row, column=4, value=s.get("seleccion",""))
        ws3.cell(row=row, column=5, value=s.get("casa",""))
        ws3.cell(row=row, column=6, value=cuota).number_format = "0.00"
        ws3.cell(row=row, column=7, value=stake).number_format = "0.00"
        rc = ws3.cell(row=row, column=8, value=s.get("resultado",""))
        rc.font = Font(name=FONT, size=10, bold=True,
                       color="1E8449" if s.get("resultado")=="GANADA" else "C0392B")
        ws3.cell(row=row, column=9, value=pnl_s).number_format = "+0.00;-0.00;0.00"
        for col in range(1, 10): ws3.cell(row=row, column=col).border = BORDER
        row += 1
    if row == 2:
        ws3.cell(row=2, column=1, value="(Todavía no hay apuestas liquidadas)")
        ws3.cell(row=2, column=1).font = Font(name=FONT, italic=True, color="888888")
        ws3.merge_cells("A2:I2")
    ws3.freeze_panes = "A2"
    if row > 2: ws3.auto_filter.ref = f"A1:I{row-1}"
    widths3 = [18,18,38,26,16,8,12,14,12]
    for i, w in enumerate(widths3, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ------- HOJA 4: EVOLUCIÓN BANKROLL -------
    ws4 = wb.create_sheet("Evolución bankroll")
    ws4.sheet_view.showGridLines = False
    ws4["A1"]="Fecha"; ws4["B1"]="Bankroll (€)"; ws4["C1"]="P&L (€)"
    for col in "ABC":
        ws4[f"{col}1"].font = HEADER_FONT; ws4[f"{col}1"].fill = HEADER_FILL
        ws4[f"{col}1"].alignment = Alignment(horizontal="center")
        ws4[f"{col}1"].border = BORDER

    hist = estado.get("historial", [])
    row = 2
    for h in hist:
        ws4.cell(row=row, column=1, value=h.get("fecha","")[:16].replace("T"," "))
        ws4.cell(row=row, column=2, value=float(h.get("bankroll", BANKROLL_INICIAL))).number_format = "0.00"
        ws4.cell(row=row, column=3, value=f"=B{row}-{BANKROLL_INICIAL}").number_format = "+0.00;-0.00;0.00"
        for c in "ABC": ws4[f"{c}{row}"].border = BORDER
        row += 1
    last = row - 1
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 15
    ws4.column_dimensions["C"].width = 15
    if last >= 2:
        chart = LineChart()
        chart.title = "Evolución del bankroll"
        chart.style = 12; chart.height = 10; chart.width = 22
        chart.y_axis.title = "€"; chart.x_axis.title = "Tiempo"
        data = Reference(ws4, min_col=2, min_row=1, max_row=last, max_col=2)
        cats = Reference(ws4, min_col=1, min_row=2, max_row=last)
        chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
        ws4.add_chart(chart, "E2")

    # ------- HOJA 5: POR DEPORTE -------
    ws5 = wb.create_sheet("Por deporte")
    ws5.sheet_view.showGridLines = False
    ws5["A1"]="Deporte"; ws5["B1"]="Apuestas abiertas"; ws5["C1"]="Stake total (€)"; ws5["D1"]="Edge promedio (%)"
    for col in "ABCD":
        ws5[f"{col}1"].font = HEADER_FONT; ws5[f"{col}1"].fill = HEADER_FILL
        ws5[f"{col}1"].alignment = Alignment(horizontal="center")
        ws5[f"{col}1"].border = BORDER

    agg = {}
    for p in pending:
        d = p.get("deporte","?").split("_")[0] or "otros"
        agg.setdefault(d, {"n":0, "stake":0.0, "edges":[]})
        agg[d]["n"] += 1
        try: agg[d]["stake"] += float(p.get("stake_eur",0) or 0)
        except: pass
        try: agg[d]["edges"].append(float(p.get("edge_pct",0)))
        except: pass

    row = 2
    for d, v in sorted(agg.items(), key=lambda kv: kv[1]["n"], reverse=True):
        ws5.cell(row=row, column=1, value=d)
        ws5.cell(row=row, column=2, value=v["n"])
        ws5.cell(row=row, column=3, value=round(v["stake"],2)).number_format = "0.00"
        ed = (sum(v["edges"])/len(v["edges"])) if v["edges"] else 0
        ws5.cell(row=row, column=4, value=round(ed,2)).number_format = "0.00"
        for c in "ABCD": ws5[f"{c}{row}"].border = BORDER
        row += 1
    last5 = row - 1
    ws5.column_dimensions["A"].width = 20
    ws5.column_dimensions["B"].width = 20
    ws5.column_dimensions["C"].width = 18
    ws5.column_dimensions["D"].width = 20
    if last5 >= 2:
        bar = BarChart(); bar.type="bar"
        bar.title = "Apuestas abiertas por deporte"
        bar.style = 11; bar.height = 10; bar.width = 18
        data = Reference(ws5, min_col=2, min_row=1, max_row=last5, max_col=2)
        cats = Reference(ws5, min_col=1, min_row=2, max_row=last5)
        bar.add_data(data, titles_from_data=True); bar.set_categories(cats)
        ws5.add_chart(bar, "F2")
        pie = PieChart()
        pie.title = "Distribución por deporte"
        data = Reference(ws5, min_col=2, min_row=1, max_row=last5, max_col=2)
        cats = Reference(ws5, min_col=1, min_row=2, max_row=last5)
        pie.add_data(data, titles_from_data=True); pie.set_categories(cats)
        pie.dataLabels = DataLabelList(showPercent=True)
        pie.height = 10; pie.width = 14
        ws5.add_chart(pie, "F22")

    # ------- HOJA 6: GUÍA -------
    ws6 = wb.create_sheet("Guía rápida")
    ws6.sheet_view.showGridLines = False
    ws6.merge_cells("B2:G2")
    ws6["B2"] = "Guía rápida — Qué es cada cosa"
    ws6["B2"].font = TITLE_FONT; ws6["B2"].fill = TITLE_FILL
    ws6["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[2].height = 28

    guia = [
        ("Value bet", "Apuesta donde la cuota del bookmaker es mejor que la cuota \"justa\" del mercado. A largo plazo, repetida muchas veces, es rentable."),
        ("Edge (%)", "Ventaja estadística estimada. Edge 5% = esperamos ganar ~5% sobre lo apostado a largo plazo."),
        ("Cuota justa", "Cuota teórica sin margen de la casa. Mediana de las cuotas de todas las casas (filtradas)."),
        ("Probabilidad real", "Lo que de verdad creemos que vale esa apuesta."),
        ("Kelly fraccional", "Fórmula para decidir cuánto apostar. Kelly/20 → solo céntimos por apuesta."),
        ("Stake", "Lo que apostamos en esa apuesta concreta."),
        ("Bankroll", "Dinero virtual disponible. Empiezas con 20 €."),
        ("P&L", "Profit & Loss. Lo ganado/perdido desde el principio."),
        ("ROI", "Rentabilidad sobre el dinero apostado."),
        ("Paper trading", "Simulación con dinero virtual. Cero riesgo."),
        ("MAD filter", "Filtro anti-cuotas-raras."),
        ("Rotación de APIs", "El bot tiene 2 claves. Si una se agota, pasa a la siguiente."),
    ]
    ws6["B4"]="Concepto"; ws6["C4"]="Qué significa"
    for c in "BC":
        ws6[f"{c}4"].font = HEADER_FONT; ws6[f"{c}4"].fill = HEADER_FILL
        ws6[f"{c}4"].alignment = Alignment(horizontal="center")

    r = 5
    for concepto, expl in guia:
        ws6.cell(row=r, column=2, value=concepto).font = Font(name=FONT, bold=True, size=10, color="1F3A5F")
        ws6.cell(row=r, column=3, value=expl).font = Font(name=FONT, size=10, color="333333")
        ws6.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws6.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws6.row_dimensions[r].height = 36
        ws6.cell(row=r, column=2).border = BORDER
        ws6.cell(row=r, column=3).border = BORDER
        r += 1

    ws6.column_dimensions["A"].width = 2
    ws6.column_dimensions["B"].width = 22
    for col in "CDEFG": ws6.column_dimensions[col].width = 18

    wb.save(OUT)
    print(f"[excel] OK -> {OUT}")
    return OUT


if __name__ == "__main__":
    generar()
